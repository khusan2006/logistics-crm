"""Excel (.xlsx) export helpers built on openpyxl.

A sheet built here is meant to be READ, not repaired: money lands as a number with a
thousands separator and two decimals, kg with three, a sana as a real date in
kk.oo.yyyy, and every column is wide enough for what is in it. The header row is bold
and frozen, so scrolling a year of to'lovlar keeps the column names in sight.

Formats are inferred from the values themselves — a Decimal is money, a date is a date
— because the alternative (a dict of column indexes at every call site) is a list that
silently stops matching its headers the first time a column is inserted. Where the
guess is wrong, `formats` overrides it BY HEADER TEXT, which cannot drift:

    xlsx_response("ombor.xlsx", headers, rows, formats={"Kg": KG})
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MONEY = "#,##0.00"
KG = "#,##0.000"
DATE = "DD.MM.YYYY"
PERCENT = "0.##"

# A column never gets narrower than its header or wider than this: one long izoh
# should not push the money columns off the screen.
_MIN_WIDTH = 9
_MAX_WIDTH = 42


def _guess_format(value):
    """What this kind of value should look like in Excel, or None to leave it alone."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (datetime, date)):
        return DATE
    if isinstance(value, (Decimal, float)):
        return MONEY
    return None


def _fill_sheet(ws, headers, rows, formats=None):
    """Write one sheet: bold frozen header, then the rows, formatted and sized."""
    formats = formats or {}
    headers = list(headers)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    widths = [len(str(h)) for h in headers]
    # Set once, off the first row that actually has a value in that column: an
    # explicit format wins, otherwise the value's own kind decides.
    column_format = [formats.get(h) for h in headers]
    decided = [fmt is not None for fmt in column_format]

    for row in rows:
        values = list(row)
        ws.append(values)
        written = ws[ws.max_row]
        for i, value in enumerate(values):
            if i >= len(widths):     # a row longer than its headers: still sized
                widths.append(0)
                column_format.append(None)
                decided.append(False)
            if value is None or value == "":
                continue
            if not decided[i]:
                column_format[i] = _guess_format(value)
                decided[i] = True
            if column_format[i]:
                written[i].number_format = column_format[i]
            shown = value.strftime("%d.%m.%Y") if isinstance(value, (datetime, date)) else str(value)
            widths[i] = max(widths[i], len(shown))

    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(
            max(width + 2, _MIN_WIDTH), _MAX_WIDTH)
    return ws


def _download(wb, filename):
    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def xlsx_response(filename, headers, rows, sheet_title=None, formats=None):
    """A one-sheet .xlsx download. `rows` is an iterable of iterables aligned with
    `headers`; money values must arrive as raw Decimals, not formatted strings —
    a string that looks like a number cannot be summed in Excel."""
    wb = Workbook()
    if sheet_title:
        wb.active.title = sheet_title
    _fill_sheet(wb.active, headers, rows, formats)
    return _download(wb, filename)


def xlsx_book_response(filename, sheets):
    """A multi-sheet .xlsx download — one tab per (title, headers, rows[, formats]).

    Two ledgers that are read against each other (the kassa's Kirim and Chiqim)
    belong in one file the reader opens once, not in two downloads to line up by
    hand."""
    wb = Workbook()
    for index, sheet in enumerate(sheets):
        title, headers, rows, *rest = sheet
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title
        _fill_sheet(ws, headers, rows, rest[0] if rest else None)
    return _download(wb, filename)
