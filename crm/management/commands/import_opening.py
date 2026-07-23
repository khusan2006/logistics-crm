"""Excel opening-balance import.

Seeds Partners, Customers, and open Contracts from an .xlsx file matching the
documented sheet layout (see docs/import-format.md). This is opening-balance
seeding, not a full history import — it is intentionally simple and idempotent
(safe to re-run the same file: matching rows are found by natural key via
get_or_create, never duplicated).

IMPORTANT — the client's real Excel files are not available yet. This command
is built against a format WE defined (docs/import-format.md). When the client's
actual files arrive, compare against that doc and adapt the column mapping if
their layout differs.

Usage:
    python manage.py import_opening path/to/opening.xlsx

Sheets (all optional — a missing sheet is skipped, not an error):
    Partners             name, phone, city, note
    Customers            name, phone, address, note
    Contracts            partner, brand, kg, price, created, deadline, note
    CustomerOpeningDebt  customer, amount   (reported only — see doc for why)

Bad rows (e.g. a blank required field) are skipped with a logged warning; the
rest of the sheet still imports.
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from crm.models import Contract, ContractLine, Customer, Partner


def _norm_header(value):
    return str(value).strip().lower() if value is not None else ""


def _clean(value):
    """Strip strings; pass through everything else (numbers, dates, None)."""
    if isinstance(value, str):
        return value.strip()
    return value


def _to_date(value, row_num, field):
    """Accept a datetime/date cell (normal openpyxl case) or an ISO string."""
    if value is None or value == "":
        return None, f"row {row_num}: {field} bo'sh"
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    try:
        return date.fromisoformat(str(value).strip()), None
    except ValueError:
        return None, f"row {row_num}: {field} sanasi noto'g'ri ({value!r})"


def _to_decimal(value, row_num, field):
    if value is None or value == "":
        return None, f"row {row_num}: {field} bo'sh"
    try:
        return Decimal(str(value)), None
    except InvalidOperation:
        return None, f"row {row_num}: {field} raqam emas ({value!r})"


def _iter_rows(worksheet):
    """Yield (row_num, {header: value}) for every non-empty data row, mapping
    columns by header name (case-insensitive), read from row 1."""
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return
    headers = [_norm_header(h) for h in header_row]
    for row_num, raw_row in enumerate(rows, start=2):
        if raw_row is None or all(v is None or v == "" for v in raw_row):
            continue
        record = {}
        for header, value in zip(headers, raw_row):
            if header:
                record[header] = _clean(value)
        yield row_num, record


class Command(BaseCommand):
    help = (
        "Import opening balances (Partners/Customers/Contracts/CustomerOpeningDebt) "
        "from a documented .xlsx layout. See docs/import-format.md."
    )

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx opening-balance file")

    def handle(self, *args, **options):
        path = options["path"]
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fayl topilmadi: {path}")
        except Exception as exc:  # openpyxl raises various zip/xml errors on bad files
            raise CommandError(f"Faylni ochib bo'lmadi: {exc}")

        sheet_names = {name.lower(): name for name in workbook.sheetnames}

        self.stdout.write(f"Import boshlandi: {path}")

        self._import_partners(workbook, sheet_names)
        self._import_customers(workbook, sheet_names)
        self._import_contracts(workbook, sheet_names)
        self._report_opening_debts(workbook, sheet_names)

        self.stdout.write(self.style.SUCCESS("Import tugadi."))

    # -- Partners ----------------------------------------------------------

    def _import_partners(self, workbook, sheet_names):
        sheet_key = sheet_names.get("partners")
        if sheet_key is None:
            self.stdout.write("Partners varag'i topilmadi — o'tkazib yuborildi")
            return

        created = updated = skipped = 0
        for row_num, row in _iter_rows(workbook[sheet_key]):
            name = row.get("name")
            if not name:
                self.stdout.write(self.style.WARNING(
                    f"  Partners row {row_num}: name bo'sh — o'tkazib yuborildi"))
                skipped += 1
                continue

            partner, was_created = Partner.objects.get_or_create(
                name=name,
                defaults={
                    "phone": row.get("phone") or "",
                    "city": row.get("city") or "",
                    "note": row.get("note") or "",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            f"Partners: {created} yaratildi, {updated} mavjud edi, {skipped} o'tkazib yuborildi")

    # -- Customers -----------------------------------------------------------

    def _import_customers(self, workbook, sheet_names):
        sheet_key = sheet_names.get("customers")
        if sheet_key is None:
            self.stdout.write("Customers varag'i topilmadi — o'tkazib yuborildi")
            return

        created = updated = skipped = 0
        for row_num, row in _iter_rows(workbook[sheet_key]):
            name = row.get("name")
            if not name:
                self.stdout.write(self.style.WARNING(
                    f"  Customers row {row_num}: name bo'sh — o'tkazib yuborildi"))
                skipped += 1
                continue

            customer, was_created = Customer.objects.get_or_create(
                name=name,
                defaults={
                    "phone": row.get("phone") or "",
                    "address": row.get("address") or "",
                    "note": row.get("note") or "",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            f"Customers: {created} yaratildi, {updated} mavjud edi, {skipped} o'tkazib yuborildi")

    # -- Contracts -------------------------------------------------------

    def _import_contracts(self, workbook, sheet_names):
        sheet_key = sheet_names.get("contracts")
        if sheet_key is None:
            self.stdout.write("Contracts varag'i topilmadi — o'tkazib yuborildi")
            return

        created = updated = skipped = 0
        for row_num, row in _iter_rows(workbook[sheet_key]):
            partner_name = row.get("partner")
            brand = row.get("brand")
            if not partner_name or not brand:
                self.stdout.write(self.style.WARNING(
                    f"  Contracts row {row_num}: partner/brand bo'sh — o'tkazib yuborildi"))
                skipped += 1
                continue

            kg, kg_error = _to_decimal(row.get("kg"), row_num, "kg")
            price, price_error = _to_decimal(row.get("price"), row_num, "price")
            created_date, created_error = _to_date(row.get("created"), row_num, "created")
            deadline_date, deadline_error = _to_date(row.get("deadline"), row_num, "deadline")

            error = kg_error or price_error or created_error or deadline_error
            if error:
                self.stdout.write(self.style.WARNING(f"  Contracts {error} — o'tkazib yuborildi"))
                skipped += 1
                continue

            partner, _ = Partner.objects.get_or_create(name=partner_name)

            contract, was_created = Contract.objects.get_or_create(
                partner=partner, lines__brand=brand, created=created_date,
                defaults={
                    "deadline": deadline_date,
                    "note": row.get("note") or "",
                },
            )
            if was_created:
                ContractLine.objects.create(
                    contract=contract, brand=brand, kg=kg, price=price)
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            f"Contracts: {created} yaratildi, {updated} mavjud edi, {skipped} o'tkazib yuborildi")

    # -- CustomerOpeningDebt (report-only) ----------------------------------

    def _report_opening_debts(self, workbook, sheet_names):
        sheet_key = sheet_names.get("customeropeningdebt")
        if sheet_key is None:
            self.stdout.write("CustomerOpeningDebt varag'i topilmadi — o'tkazib yuborildi")
            return

        reported = skipped = 0
        self.stdout.write(self.style.WARNING(
            "CustomerOpeningDebt: qarzlar avtomatik yozilmaydi (sotuvsiz qarz "
            "obyekti yo'q) — quyidagilarni qo'lda kiriting:"))
        for row_num, row in _iter_rows(workbook[sheet_key]):
            customer_name = row.get("customer")
            amount, amount_error = _to_decimal(row.get("amount"), row_num, "amount")
            if not customer_name or amount_error:
                self.stdout.write(self.style.WARNING(
                    f"  CustomerOpeningDebt row {row_num}: noto'g'ri qator — o'tkazib yuborildi"))
                skipped += 1
                continue

            self.stdout.write(f"  qo'lda kiritish kerak: {customer_name} — {amount}")
            reported += 1

        self.stdout.write(
            f"CustomerOpeningDebt: {reported} qator ro'yxatga olindi (qo'lda kiritish uchun), "
            f"{skipped} o'tkazib yuborildi")
