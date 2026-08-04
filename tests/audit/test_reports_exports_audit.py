"""Audit pass — Hisobotlar (crm/views.py:2318) and the five .xlsx exports.

TRIAGED. The first tester's run was cut off before it could check its own xfail
markers against the source, so every marker here has now been re-read against
crm/views.py, crm/models.py and crm/forms.py and either

  * kept and SHARPENED — the code really is wrong, the reason names file:line and
    the concrete wrong figure, or
  * WITHDRAWN — the code was doing exactly what its own comment says it does, so
    the test now asserts that intended behaviour and a `# withdrawn:` comment says
    what the original claim got wrong.

Diagnosis only: nothing outside tests/audit/ is touched.

An export is the highest-stakes place for a currency mix: the number leaves the
app and is read in Excel with no toggle, no tooltip and no row context. So every
probe here asks the same four questions —

    is the TYPED side of a dual-stored row the one that reaches the file,
    is the file's column total the same number the screen printed,
    does a filter narrow the screen and the file identically,
    and is the cell a NUMBER Excel can sum rather than a rendered string.

Run:
    TEST_DB_SUFFIX=_trireports_exports .venv/bin/python -m pytest \
        tests/audit/test_reports_exports_audit.py -q -p no:randomly
"""
from decimal import Decimal
from html.parser import HTMLParser
from io import BytesIO
from unittest.mock import patch

import openpyxl
import pytest

from conftest import line_data
from crm.models import (
    Contract, ContractLine, Currency, Customer, CustomerPayment, Partner, Sale,
    Shipment, ShipmentLine, ShipmentStatus, SupplierPayment,
)

CONTRACTS = "/reports/export/contracts.xlsx"
SUP_PAYS = "/reports/export/supplier-payments.xlsx"
SHIPMENTS = "/reports/export/shipments.xlsx"
SALES = "/reports/export/sales.xlsx"
DEBTS = "/reports/export/debts.xlsx"
ALL_EXPORTS = (CONTRACTS, SUP_PAYS, SHIPMENTS, SALES, DEBTS)


# ── workbook helpers ─────────────────────────────────────────────────────────

def _sheet(resp):
    assert resp.status_code == 200, resp.status_code
    return openpyxl.load_workbook(BytesIO(resp.content)).active


def _header(ws):
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _rows(resp):
    """The data rows as dicts keyed by the header cell above them."""
    ws = _sheet(resp)
    head = _header(ws)
    return [dict(zip(head, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]


def _cells(resp):
    """Same shape as _rows but keeping the openpyxl Cell, so a test can look at
    data_type — a figure Excel cannot sum is stored as 's', not 'n'."""
    ws = _sheet(resp)
    head = _header(ws)
    return [dict(zip(head, list(row))) for row in ws.iter_rows(min_row=2)]


def _dec(value):
    return Decimal(str(value if value is not None else 0))


def _total(rows, column):
    return sum((_dec(r[column]) for r in rows), Decimal("0"))


def _close(left, right, quantum="0.01"):
    """Excel round-trips a Decimal through a float, so a column total is compared
    to the screen's Decimal at the column's own precision, not bit for bit."""
    return abs(_dec(left) - _dec(right)) <= Decimal(quantum)


def _once_per(rows, key_column, value_column):
    """Total a column that repeats down the rows of one parent — the per-kelishuv
    money in kelishuvlar.xlsx — by counting each parent exactly once."""
    seen = {}
    for row in rows:
        seen.setdefault(row[key_column], _dec(row[value_column]))
    return sum(seen.values(), Decimal("0"))


def _status_or_error(client, url, params):
    """The HTTP status, or the exception class name when the view let one escape.

    The test client re-raises whatever the view raised; in production the same
    escape is a 500. Collecting instead of raising lets ONE test report every URL
    that breaks rather than stopping at the first."""
    try:
        return client.get(url, params).status_code
    except Exception as exc:
        return type(exc).__name__


# ── data helpers ─────────────────────────────────────────────────────────────

def _partner(name="Pars"):
    return Partner.objects.create(name=name, phone="1", city="Tehron")


def _line(contract, brand="LLDPE", kg="10000", price="1.00",
          currency="usd", rate="12000", price_uzs=None):
    if price_uzs is None:
        price_uzs = Decimal(price) * Decimal(rate)
    return ContractLine.objects.create(
        contract=contract, brand=brand, kg=Decimal(kg), price=Decimal(price),
        price_uzs=Decimal(price_uzs), currency=currency,
        exchange_rate=Decimal(rate))


def _contract(partner=None, created="2026-07-01", **line_kw):
    contract = Contract.objects.create(partner=partner or _partner(), created=created)
    _line(contract, **line_kw)
    return contract


def _lot(contract, kg="5000", eta="2026-07-15", arrived="2026-07-16"):
    shipment = Shipment.objects.create(
        contract=contract, status=ShipmentStatus.arrival(), sent="2026-07-05",
        eta=eta, arrived=arrived, transport="01A111AA", container="MSCU-1")
    return ShipmentLine.objects.create(
        shipment=shipment, contract_line=contract.lines.first(), kg=Decimal(kg))


def _customer(name="Alisher Mebel"):
    return Customer.objects.create(name=name, phone="998901112233", address="Toshkent")


def _sale(customer, lot, kg="100", price="2.00", currency="usd", rate="12000",
          price_uzs=None, date_="2026-07-17"):
    if price_uzs is None:
        price_uzs = Decimal(price) * Decimal(rate)
    return Sale.objects.create(
        customer=customer, line=lot, kg=Decimal(kg), price=Decimal(price),
        price_uzs=Decimal(price_uzs), currency=currency,
        exchange_rate=Decimal(rate), date=date_)


class _FormScraper(HTMLParser):
    """Reads a rendered modal the way a browser does, so the payload below is
    literally what pressing Saqlash sends — not what the Python form object thinks
    it would send. The base.html currency enhancer only appends a preview <span>
    (templates/base.html:1288); it never rewrites an input value, so the server-
    rendered value IS the value that gets posted back."""

    def __init__(self):
        super().__init__()
        self.data, self._select, self._textarea = {}, None, None

    def handle_starttag(self, tag, attrs):
        attr = dict(attrs)
        name = attr.get("name")
        if tag == "input" and name:
            kind = (attr.get("type") or "text").lower()
            if kind in ("checkbox", "radio"):
                if "checked" in attr:
                    self.data[name] = attr.get("value", "on")
                else:
                    self.data.setdefault(name, "")
            elif kind != "submit":
                self.data[name] = attr.get("value", "")
        elif tag == "select" and name:
            self._select = name
            self.data.setdefault(name, "")
        elif tag == "option" and self._select and "selected" in attr:
            self.data[self._select] = attr.get("value", "")
        elif tag == "textarea" and name:
            self._textarea = name
            self.data.setdefault(name, "")

    def handle_endtag(self, tag):
        if tag == "select":
            self._select = None
        elif tag == "textarea":
            self._textarea = None

    def handle_data(self, data):
        if self._textarea:
            self.data[self._textarea] = self.data.get(self._textarea, "") + data


def _resubmit_payload(client, url):
    """Open the edit modal and hand back the form exactly as it stands — the
    "operator opened it and pressed Saqlash without touching anything" payload."""
    resp = client.get(url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
    assert resp.status_code == 200, resp.status_code
    scraper = _FormScraper()
    scraper.feed(resp.content.decode())
    return scraper.data


# ── (a) round-trip: the typed side is the one that reaches the file ──────────

def test_a_som_kelishuv_narx_reaches_the_xlsx_as_som(admin_client, db):
    """Typed 9 768 so'm/kg at 12 000. The file must carry that exact figure in the
    so'm column and the derived 0.814 in the dollar one — never the other way."""
    partner = _partner()
    resp = admin_client.post("/contracts/new/", {
        "partner": partner.pk, "currency": "uzs", "created": "2026-07-01",
        "note": "", "planned_trucks": "1",
        **line_data({"brand": "LLDPE", "kg": "1000", "price": "9768"}),
    })
    assert resp.status_code == 302

    row = _rows(admin_client.get(CONTRACTS))[0]
    assert row["Valyuta"] == "So'm"
    assert _dec(row["Kurs"]) == Decimal("12000")
    assert _dec(row["Narx (so'm)"]) == Decimal("9768")        # typed, exact
    assert _dec(row["Narx ($)"]) == Decimal("0.814")          # derived
    assert _dec(row["Jami (so'm)"]) == Decimal("9768000")
    assert _dec(row["Jami ($)"]) == Decimal("814.00")


def test_a_som_sotuv_narx_reaches_the_xlsx_as_som(admin_client, db):
    contract = _contract(kg="10000")
    lot = _lot(contract)
    customer = _customer()
    resp = admin_client.post(f"/sales/new/?lot={lot.pk}", {
        "customer": customer.pk, "kg": "1000", "currency": "uzs",
        "price": "14040", "exchange_rate": "12000", "date": "2026-07-18",
        "debt_deadline": "", "note": "",
    })
    assert resp.status_code == 302

    row = _rows(admin_client.get(SALES))[0]
    assert row["Valyuta"] == "So'm"
    assert _dec(row["Sotuv narx (so'm)"]) == Decimal("14040")   # typed, exact
    assert _dec(row["Sotuv narx ($)"]) == Decimal("1.17")       # derived
    assert _dec(row["Jami (so'm)"]) == Decimal("14040000")
    assert _dec(row["Jami ($)"]) == Decimal("1170.00")


def test_a_som_hamkor_tolovi_reaches_the_xlsx_as_som(admin_client, db):
    contract = _contract(kg="10000", price="1.00")
    resp = admin_client.post("/supplier-payments/new/", {
        "contract": contract.pk, "date": "2026-07-02", "currency": "uzs",
        "amount": "12000000", "exchange_rate": "12000", "commission_percent": "",
        "method": "cash", "fee_percent": "0", "note": "",
    })
    assert resp.status_code == 302

    row = _rows(admin_client.get(SUP_PAYS))[0]
    assert row["Valyuta"] == "So'm"
    assert _dec(row["Hamkorga (so'm)"]) == Decimal("12000000")  # typed, exact
    assert _dec(row["Hamkorga ($)"]) == Decimal("1000.00")      # derived


def test_a_som_figure_that_does_not_divide_evenly_still_ships_exact(admin_client, db):
    """Round-trip probe with a kurs that leaves a remainder: 13 000 000 so'm at
    12 345 is $1 053.06 to the cent. convert_pair (crm/models.py:37) promises the
    typed side comes back untouched, so the so'm column must be the round figure
    and only the dollar one may carry the rounding."""
    contract = _contract(kg="10000", price="1.00")
    assert admin_client.post("/supplier-payments/new/", {
        "contract": contract.pk, "date": "2026-07-02", "currency": "uzs",
        "amount": "13000000", "exchange_rate": "12345", "commission_percent": "",
        "method": "cash", "fee_percent": "0", "note": "",
    }).status_code == 302

    row = _rows(admin_client.get(SUP_PAYS))[0]
    assert _dec(row["Hamkorga (so'm)"]) == Decimal("13000000")   # typed, undivided
    assert _dec(row["Hamkorga ($)"]) == Decimal("1053.06")       # 13e6/12345, 2dp


def test_a_dollar_row_and_a_som_row_keep_their_own_side_in_one_file(admin_client, db):
    """Two hamkor to'lovlari booked in different currencies at different kursi.
    Neither so'm figure may be a reconversion of the other row's dollar side."""
    contract = _contract(kg="100000", price="1.00")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.USD,
        amount=Decimal("1000.00"), amount_uzs=Decimal("12000000.00"),
        exchange_rate=Decimal("12000"), method="cash")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-03", currency=Currency.UZS,
        amount=Decimal("1000.00"), amount_uzs=Decimal("13500000.00"),
        exchange_rate=Decimal("13500"), method="cash")

    rows = sorted(_rows(admin_client.get(SUP_PAYS)), key=lambda r: str(r["Sana"]))
    assert [r["Valyuta"] for r in rows] == ["Dollar", "So'm"]
    assert _dec(rows[0]["Hamkorga (so'm)"]) == Decimal("12000000")
    assert _dec(rows[1]["Hamkorga (so'm)"]) == Decimal("13500000")
    # the so'm column is NOT one rate applied to the dollar column
    assert _total(rows, "Hamkorga (so'm)") == Decimal("25500000")
    assert _total(rows, "Hamkorga ($)") == Decimal("2000.00")


# ── (c) currency stickiness ──────────────────────────────────────────────────

def test_a_som_sotuv_stays_som_in_the_export_after_a_dollar_tolov(admin_client, db):
    """A dollar to'lov landing against a so'm sotuv must not convert the sotuv.
    The row's Valyuta stays So'm, its narx columns do not move, and Qoldiq (so'm)
    is the so'm total less the to'lov rated at THIS sotuv's kurs."""
    lot = _lot(_contract(kg="100000", price="1.00"), kg="50000")
    customer = _customer()
    sale = _sale(customer, lot, kg="1000", price="1.1700", currency="uzs",
                 rate="12000", price_uzs="14040")
    before = _rows(admin_client.get(SALES))[0]

    assert admin_client.post("/customer-payments/new/", {
        "customer": customer.pk, "date": "2026-07-20",
        "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "0",
        "form-MIN_NUM_FORMS": "0", "form-MAX_NUM_FORMS": "1000",
        "form-0-currency": "usd", "form-0-amount": "500",
        "form-0-exchange_rate": "13500", "form-0-method": "cash",
        "form-0-fee_percent": "0", "form-0-note": "",
    }).status_code in (302, 204)

    after = _rows(admin_client.get(SALES))[0]
    assert after["Valyuta"] == "So'm"
    assert _dec(after["Sotuv narx (so'm)"]) == _dec(before["Sotuv narx (so'm)"])
    assert _dec(after["Jami (so'm)"]) == Decimal("14040000")
    sale.refresh_from_db()
    assert _close(after["Qoldiq (so'm)"], sale.remaining_uzs)
    # rated at the SOTUV's 12 000, not the to'lov's 13 500
    assert _close(after["Qoldiq (so'm)"], Decimal("14040000") - Decimal("6000000"))


# ── numbers must stay numbers ────────────────────────────────────────────────

def test_every_money_cell_is_a_number_excel_can_sum(admin_client, db):
    """A pre-formatted "1 200,50 so'm" would sum to zero in Excel. Every money and
    kg column has to arrive as a numeric cell."""
    contract = _contract(kg="10000")
    lot = _lot(contract)
    customer = _customer()
    _sale(customer, lot)
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-11", amount=Decimal("200.00"),
        amount_uzs=Decimal("2400000.00"), method="cash")
    CustomerPayment.objects.create(
        customer=customer, date="2026-07-17", amount=Decimal("50.00"),
        amount_uzs=Decimal("600000.00"), method="cash")

    numeric = {
        CONTRACTS: ["Kg", "Kurs", "Narx ($)", "Narx (so'm)", "Jami ($)",
                    "Jami (so'm)", "Yuborilgan kg", "To'langan ($)",
                    "To'langan (so'm)", "Qarz ($)", "Qarz (so'm)"],
        SUP_PAYS: ["Kurs", "Hamkorga ($)", "Hamkorga (so'm)", "Vositachi %",
                   "Vositachi ($)", "Perechisleniya %", "Perechisleniya ($)",
                   "Kassadan ($)", "Kassadan (so'm)"],
        SHIPMENTS: ["Kg"],
        SALES: ["Kg", "Kurs", "Tan narx ($)", "Sotuv narx ($)", "Sotuv narx (so'm)",
                "Jami ($)", "Jami (so'm)", "Foyda ($)", "Foyda (so'm)",
                "Qoldiq ($)", "Qoldiq (so'm)"],
        DEBTS: ["Jami savdo ($)", "Jami savdo (so'm)", "To'langan ($)",
                "To'langan (so'm)", "Qarz ($)", "Qarz (so'm)"],
    }
    for url, columns in numeric.items():
        rows = _cells(admin_client.get(url))
        assert rows, url
        for row in rows:
            for column in columns:
                cell = row[column]
                assert cell.data_type == "n", f"{url} · {column} · {cell.value!r}"
                assert isinstance(cell.value, (int, float)), f"{url} · {column}"


def test_date_cells_are_real_dates_not_text(admin_client, db):
    contract = _contract(kg="10000")
    lot = _lot(contract)
    _sale(_customer(), lot)
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-11", amount=Decimal("200.00"),
        amount_uzs=Decimal("2400000.00"), method="cash")

    for url, column in ((CONTRACTS, "Sana"), (SUP_PAYS, "Sana"),
                        (SHIPMENTS, "Jo'natilgan"), (SALES, "Sana")):
        value = _cells(admin_client.get(url))[0][column].value
        assert hasattr(value, "year"), f"{url} · {column} · {value!r}"


# ── (d) file totals must equal the screen totals ─────────────────────────────

def test_contracts_export_jami_sums_to_the_screen_kontrakt_summasi(admin_client, db):
    """A kelishuv whose two products were agreed in different currencies at
    different kursi. Both column totals must land on the KPI exactly."""
    contract = Contract.objects.create(partner=_partner(), created="2026-07-01")
    _line(contract, brand="LLDPE", kg="1000", price="1.00", currency="usd", rate="12000")
    _line(contract, brand="HDPE", kg="2000", price="0.8140", currency="uzs",
          rate="12000", price_uzs="9768")

    rows = _rows(admin_client.get(CONTRACTS))
    screen = admin_client.get("/reports/").context
    assert len(rows) == 2
    assert _total(rows, "Jami ($)") == screen["kontrakt_summasi"]
    assert _total(rows, "Jami (so'm)") == screen["kontrakt_summasi_uzs"]


# withdrawn: the original marker called the repeated To'langan/Qarz a bug ("summing
# the column gives 2x the KPI"). It is not — crm/views.py:2412-2415 states the
# decision in as many words: "One row per product, so a multi-product kelishuv is
# readable in Excel. The money columns are per kelishuv, so they repeat down its
# rows." Narx/Jami are per PRODUCT, To'langan/Qarz are per KELISHUV, and that is
# deliberate. What the file owes the reader is that the repeated value be the
# kelishuv's real figure and reconcile with the screen once per kelishuv — which is
# what this now asserts. (The readability complaint — nothing in the file marks the
# column non-additive — is a design opinion, not a defect.)
def test_contracts_export_repeats_the_per_kelishuv_money_on_every_product_row(admin_client, db):
    contract = Contract.objects.create(partner=_partner(), created="2026-07-01")
    lldpe = _line(contract, brand="LLDPE", kg="1000", price="1.00")
    hdpe = _line(contract, brand="HDPE", kg="2000", price="1.00")
    # both products shipped, so Contract.debt (shipped_value - paid) is the whole
    # kelishuv less the to'lov rather than an avans
    shipment = Shipment.objects.create(
        contract=contract, status=ShipmentStatus.arrival(), sent="2026-07-05",
        eta="2026-07-15", arrived="2026-07-16", transport="01A111AA", container="MSCU-1")
    ShipmentLine.objects.create(shipment=shipment, contract_line=lldpe, kg=Decimal("1000"))
    ShipmentLine.objects.create(shipment=shipment, contract_line=hdpe, kg=Decimal("2000"))
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-11", amount=Decimal("500.00"),
        amount_uzs=Decimal("6000000.00"), method="cash")

    rows = _rows(admin_client.get(CONTRACTS))
    screen = admin_client.get("/reports/").context
    assert len(rows) == 2
    assert [r["Marka"] for r in rows] == ["LLDPE", "HDPE"]
    # per-PRODUCT columns differ down the rows
    assert [_dec(r["Jami ($)"]) for r in rows] == [Decimal("1000"), Decimal("2000")]
    # per-KELISHUV columns repeat, each carrying the kelishuv's own figure
    assert {_dec(r["To'langan ($)"]) for r in rows} == {Decimal("500")}
    assert {_dec(r["Qarz ($)"]) for r in rows} == {Decimal("2500")}
    assert {_dec(r["To'langan (so'm)"]) for r in rows} == {Decimal("6000000")}
    # counted once per kelishuv they are exactly the screen's KPIs
    assert _once_per(rows, "Kelishuv", "To'langan ($)") == screen["hamkorga_tolangan"]
    assert _once_per(rows, "Kelishuv", "To'langan (so'm)") == screen["hamkorga_tolangan_uzs"]
    assert _once_per(rows, "Kelishuv", "Qarz ($)") == screen["hamkor_qarzi"]
    assert _once_per(rows, "Kelishuv", "Qarz (so'm)") == screen["hamkor_qarzi_uzs"]


def test_supplier_payments_export_sums_to_the_screen_kpi(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.USD,
        amount=Decimal("1000.00"), amount_uzs=Decimal("12000000.00"),
        exchange_rate=Decimal("12000"), method="cash")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-03", currency=Currency.UZS,
        amount=Decimal("740.74"), amount_uzs=Decimal("10000000.00"),
        exchange_rate=Decimal("13500"), method="cash")

    rows = _rows(admin_client.get(SUP_PAYS))
    screen = admin_client.get("/reports/").context
    assert _total(rows, "Hamkorga ($)") == screen["hamkorga_tolangan"]
    assert _total(rows, "Hamkorga (so'm)") == screen["hamkorga_tolangan_uzs"]


def test_sales_export_foyda_sums_to_the_screen_profit_total(admin_client, db):
    """Two sotuvlar, one struck in dollars and one in so'm at another kurs — the
    hardest case for a profit column, since profit has no stored so'm twin."""
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    customer = _customer()
    _sale(customer, lot, kg="1000", price="2.00", currency="usd", rate="12000")
    _sale(customer, lot, kg="1000", price="1.5000", currency="uzs",
          rate="13500", price_uzs="20250")

    rows = _rows(admin_client.get(SALES))
    screen = admin_client.get("/reports/").context
    assert len(rows) == 2
    assert _close(_total(rows, "Foyda ($)"), screen["profit_total"])
    assert _close(_total(rows, "Foyda (so'm)"), screen["profit_total_uzs"])


def test_debts_export_qarz_sums_to_the_screen_mijoz_qarzi(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    first, second = _customer("Alisher"), _customer("Bobur")
    _sale(first, lot, kg="1000", price="2.00")
    _sale(second, lot, kg="500", price="1.5000", currency="uzs",
          rate="13500", price_uzs="20250")
    CustomerPayment.objects.create(
        customer=first, date="2026-07-18", amount=Decimal("500.00"),
        amount_uzs=Decimal("6000000.00"), method="cash")

    rows = _rows(admin_client.get(DEBTS))
    screen = admin_client.get("/reports/").context
    assert len(rows) == 2
    assert _close(_total(rows, "Qarz ($)"), screen["mijoz_qarzi"])
    assert _close(_total(rows, "Qarz (so'm)"), screen["mijoz_qarzi_uzs"])


def test_shipments_export_kg_matches_the_screen_yuborilgan_kg(admin_client, db):
    """Aggregate consistency for the one export with no money in it: the Kg column
    is the Yuborilgan kg KPI, and a Holat filter must move both together."""
    contract = _contract(kg="100000", price="1.00")
    _lot(contract, kg="20000")
    _lot(contract, kg="30000")

    rows = _rows(admin_client.get(SHIPMENTS))
    screen = admin_client.get("/reports/").context
    assert _total(rows, "Kg") == _dec(screen["yuborilgan_kg"]) == Decimal("50000")

    status = ShipmentStatus.objects.exclude(pk=ShipmentStatus.arrival().pk).first()
    q = {"status": status.pk}
    assert _rows(admin_client.get(SHIPMENTS, q)) == []
    assert _dec(admin_client.get("/reports/", q).context["yuborilgan_kg"]) == Decimal("0")


# ── (d) a filter has to narrow the screen and the file identically ──────────

def test_the_date_filter_narrows_the_sales_export_like_the_screen(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    customer = _customer()
    _sale(customer, lot, kg="1000", price="2.00", date_="2026-05-01")
    _sale(customer, lot, kg="1000", price="2.00", date_="2026-07-17")

    window = {"from": "2026-07-01", "to": "2026-07-31"}
    rows = _rows(admin_client.get(SALES, window))
    screen = admin_client.get("/reports/", window).context
    assert len(rows) == 1
    assert str(rows[0]["Sana"])[:10] == "2026-07-17"
    assert _close(_total(rows, "Foyda ($)"), screen["profit_total"])


@pytest.mark.xfail(reason="UPHELD. crm/views.py:2476 export_debts is the only one of the "
                          "five exports that never calls _report_querysets — its rows come "
                          "straight from `Customer.objects.all() if c.balance > 0` "
                          "(crm/views.py:2482), so it silently drops the querystring that "
                          "reports.html:42 hands it. Report filtered to 2026-07-01..07-31, "
                          "whose only sotuv is dated 2026-03-05: the screen's Mijozlar "
                          "bo'yicha table is empty, qarzdorlar.xlsx still exports 'Mart "
                          "mijozi' with Jami savdo $2 000 / 24 000 000 so'm. Nothing in the "
                          "file says the period was dropped. Fix: build the rows from "
                          "_report_querysets(request)['sales'] the way reports() builds "
                          "customer_rows (crm/views.py:2370), or drop the querystring from "
                          "the Qarzdorlar link and label the file 'jami'.",
                   strict=False)
def test_the_date_filter_reaches_the_qarzdorlar_export(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    _sale(_customer("Mart mijozi"), lot, kg="1000", price="2.00", date_="2026-03-05")

    window = {"from": "2026-07-01", "to": "2026-07-31"}
    screen = admin_client.get("/reports/", window).context
    rows = _rows(admin_client.get(DEBTS, window))
    assert screen["customer_rows"] == []      # the March sotuv is out of the window
    assert rows == []                         # so the file must be empty too


@pytest.mark.xfail(reason="UPHELD — same root cause as the date case, second filter. "
                          "crm/views.py:2476 export_debts ignores ?partner= too. With two "
                          "hamkorlar each owed by one mijoz, ?partner=<Pars> narrows the "
                          "screen's Mijozlar bo'yicha table to ['Pars mijozi'] while "
                          "qarzdorlar.xlsx exports ['Kaveh mijozi', 'Pars mijozi'] — a "
                          "debtor who never bought a gram of that hamkor's goods, at $2 000. "
                          "Fix: same as above, one call to _report_querysets covers both "
                          "filters.",
                   strict=False)
def test_the_partner_filter_reaches_the_qarzdorlar_export(admin_client, db):
    pars, kaveh = _partner("Pars"), _partner("Kaveh")
    lot_a = _lot(_contract(partner=pars, kg="100000", price="1.00"), kg="50000")
    lot_b = _lot(_contract(partner=kaveh, kg="100000", price="1.00"), kg="50000")
    _sale(_customer("Pars mijozi"), lot_a, kg="1000", price="2.00")
    _sale(_customer("Kaveh mijozi"), lot_b, kg="1000", price="2.00")

    q = {"partner": pars.pk}
    screen = admin_client.get("/reports/", q).context
    rows = _rows(admin_client.get(DEBTS, q))
    assert [r["customer"].name for r in screen["customer_rows"]] == ["Pars mijozi"]
    assert [r["Mijoz"] for r in rows] == ["Pars mijozi"]


@pytest.mark.xfail(reason="UPHELD. crm/views.py:2273 applies Marka to CONTRACTS "
                          "(`contracts.filter(lines__brand=brand).distinct()`) and nothing "
                          "narrows the lines afterwards, while export_contracts "
                          "(crm/views.py:2421) walks `c.lines.all()`. A kelishuv holding "
                          "LLDPE 1 000 kg @ $1 and HDPE 2 000 kg @ $5 exports BOTH rows "
                          "under ?brand=LLDPE — kelishuvlar.xlsx carries a Marka column "
                          "reading HDPE under a filter that excluded HDPE. Fix: pass the "
                          "brand through and iterate `c.lines.filter(brand=brand)` when it "
                          "is set (same for export_shipments, crm/views.py:2444).",
                   strict=False)
def test_the_marka_filter_narrows_the_kelishuvlar_export_to_that_marka(admin_client, db):
    contract = Contract.objects.create(partner=_partner(), created="2026-07-01")
    _line(contract, brand="LLDPE", kg="1000", price="1.00")
    _line(contract, brand="HDPE", kg="2000", price="5.00")

    rows = _rows(admin_client.get(CONTRACTS, {"brand": "LLDPE"}))
    assert [r["Marka"] for r in rows] == ["LLDPE"]
    assert _total(rows, "Jami ($)") == Decimal("1000.00")


@pytest.mark.xfail(reason="UPHELD — the screen half of the same crm/views.py:2273 defect, "
                          "and the proof it is an oversight rather than a decision: ONE "
                          "Marka dropdown is read two different ways on one page. Sales are "
                          "filtered per LINE (crm/views.py:2287 "
                          "`sales.filter(line__contract_line__brand=brand)`) but "
                          "kelishilgan_kg (crm/views.py:2332) and kontrakt_summasi sum every "
                          "line of the matching kelishuvlar. With LLDPE 1 000 kg @ $1 beside "
                          "HDPE 2 000 kg @ $5, ?brand=LLDPE reports Kelishilgan kg 3 000 and "
                          "Kontrakt summasi $11 000 — 11x the LLDPE figure — next to a "
                          "Sotuvdan foyda that counts LLDPE only.",
                   strict=False)
def test_the_marka_filter_narrows_the_screen_kpis_to_that_marka(admin_client, db):
    contract = Contract.objects.create(partner=_partner(), created="2026-07-01")
    _line(contract, brand="LLDPE", kg="1000", price="1.00")
    _line(contract, brand="HDPE", kg="2000", price="5.00")

    screen = admin_client.get("/reports/", {"brand": "LLDPE"}).context
    assert _dec(screen["kelishilgan_kg"]) == Decimal("1000")
    assert screen["kontrakt_summasi"] == Decimal("1000.00")


# ── the screen's own rows have to add up ─────────────────────────────────────

@pytest.mark.xfail(reason="UPHELD. crm/views.py:2377 the Mijozlar bo'yicha table mixes two "
                          "periods inside one row. Sotildi/To'landi are summed from the "
                          "DATE-FILTERED sales/cust_pays (crm/views.py:2372-2375) but Qarz "
                          "is `customer.balance` — every sotuv and to'lov ever. One mijoz, "
                          "$1 000 sold in May and $500 in July, filtered to July: the row "
                          "reads Sotildi $500 · To'landi $0 · Qarz $1 500. The line above it "
                          "in the same view (crm/views.py:2371) even says the sums are net "
                          "'so the row reconciles with the net-based qarz column' — under a "
                          "date filter it cannot. The KPI beside it is labelled '(jami)' "
                          "(reports.html:75); this column is not. Fix: either sum the qarz "
                          "from the filtered c_sales/cust_pays, or label the column "
                          "Qarz (jami).",
                   strict=False)
def test_customer_row_qarz_equals_sotildi_minus_tolandi_under_a_date_filter(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    customer = _customer()
    _sale(customer, lot, kg="500", price="2.00", date_="2026-05-01")   # out of window
    _sale(customer, lot, kg="250", price="2.00", date_="2026-07-17")   # in window

    screen = admin_client.get(
        "/reports/", {"from": "2026-07-01", "to": "2026-07-31"}).context
    row = screen["customer_rows"][0]
    assert row["sotildi"] == Decimal("500.00")
    assert row["tolandi"] == Decimal("0")
    assert row["qarz"] == row["sotildi"] - row["tolandi"]


def test_customer_row_qarz_adds_up_with_no_date_filter(admin_client, db):
    """The control for the test above: with no window the same three numbers do
    reconcile, which pins the defect to the date filter rather than to net-vs-gross
    or to the balance property."""
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    customer = _customer()
    _sale(customer, lot, kg="500", price="2.00", date_="2026-05-01")
    _sale(customer, lot, kg="250", price="2.00", date_="2026-07-17")
    CustomerPayment.objects.create(
        customer=customer, date="2026-07-18", amount=Decimal("400.00"),
        amount_uzs=Decimal("4800000.00"), method="cash")

    row = admin_client.get("/reports/").context["customer_rows"][0]
    assert row["sotildi"] == Decimal("1500.00")
    assert row["tolandi"] == Decimal("400.00")
    assert row["qarz"] == row["sotildi"] - row["tolandi"] == Decimal("1100.00")


# ── every money column is supposed to ship both currencies ───────────────────

@pytest.mark.xfail(reason="UPHELD. crm/views.py:2462 the Sotuvlar export ships 'Tan narx "
                          "($)' with no so'm twin, though `Sale.cost_price_uzs` "
                          "(crm/models.py:1373) exists, is what the sotuv screens draw, and "
                          "every other money column in this very file ships both sides — the "
                          "sibling export states the rule outright at crm/views.py:2414: "
                          "'Both currencies ship in every export'. On the so'm sotuv below "
                          "the reader gets Sotuv narx (so'm) 20 250 beside Tan narx ($) "
                          "1.0000, so the obvious Excel margin column subtracts dollars from "
                          "so'm and reports a 20 249 so'm/kg margin on a 6 750 one. Fix: add "
                          "'Tan narx (so'm)' carrying s.cost_price_uzs.",
                   strict=False)
def test_sales_export_ships_a_som_twin_for_every_money_column(admin_client, db):
    lot = _lot(_contract(kg="100000", price="1.00"), kg="50000")
    _sale(_customer(), lot, kg="1000", price="1.5000", currency="uzs",
          rate="13500", price_uzs="20250")

    head = _header(_sheet(admin_client.get(SALES)))
    dollar_only = [h for h in head
                   if h.endswith("($)") and h.replace("($)", "(so'm)") not in head]
    assert dollar_only == []


@pytest.mark.xfail(reason="UPHELD. crm/views.py:2430 the Hamkor to'lovlari export ships "
                          "'Vositachi ($)' and 'Perechisleniya ($)' with no so'm twin, "
                          "though SupplierPayment.commission_amount_uzs / fee_amount_uzs "
                          "(crm/models.py:694, 698) exist and Kassadan (so'm) already "
                          "includes them (crm/models.py:703). The file therefore cannot be "
                          "reconciled in its own so'm column: the to'lov below exports "
                          "Hamkorga 13 500 000 so'm and Kassadan 13 905 000 so'm with the "
                          "405 000 so'm difference appearing nowhere — the two charges that "
                          "make it up are dollars only. Fix: add \"Vositachi (so'm)\" and "
                          "\"Perechisleniya (so'm)\".",
                   strict=False)
def test_supplier_payment_export_ships_a_som_twin_for_every_money_column(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.UZS,
        amount=Decimal("1000.00"), amount_uzs=Decimal("13500000.00"),
        exchange_rate=Decimal("13500"), commission_percent=Decimal("2"),
        method="transfer", fee_percent=Decimal("1"))

    head = _header(_sheet(admin_client.get(SUP_PAYS)))
    dollar_only = [h for h in head
                   if h.endswith("($)") and h.replace("($)", "(so'm)") not in head]
    assert dollar_only == []
    # and what the missing columns cost the reader: 405 000 so'm of the file's own
    # arithmetic is unshown, so this reconciliation cannot be written in Excel today
    row = _rows(admin_client.get(SUP_PAYS))[0]
    gap = _dec(row["Kassadan (so'm)"]) - _dec(row["Hamkorga (so'm)"])
    assert gap == Decimal("405000")
    assert gap == _dec(row.get("Vositachi (so'm)")) + _dec(row.get("Perechisleniya (so'm)"))


def test_kassadan_som_equals_its_parts_in_the_export(admin_client, db):
    """Whatever the file prints as Kassadan (so'm) has to be the sum of the so'm
    worth of its three parts at the row's own kurs."""
    contract = _contract(kg="100000", price="1.00")
    payment = SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.UZS,
        amount=Decimal("1000.00"), amount_uzs=Decimal("13500000.00"),
        exchange_rate=Decimal("13500"), commission_percent=Decimal("2"),
        method="transfer", fee_percent=Decimal("1"))

    row = _rows(admin_client.get(SUP_PAYS))[0]
    expected = (payment.amount_uzs + payment.commission_amount_uzs
                + payment.fee_amount_uzs)
    assert _close(row["Kassadan (so'm)"], expected)
    assert _close(row["Kassadan ($)"], payment.total_out)


# ── (b) idempotence: re-saving unchanged must not move a figure ─────────────

@pytest.mark.xfail(reason="UPHELD — the known MoneyEntryFormMixin seed-the-typed-side defect "
                          "(tests/audit/test_som_edit_dataloss.py) reaching the exports. "
                          "SupplierPaymentForm (crm/forms.py:658) binds `amount` straight "
                          "off the model's USD column while `currency` still reads So'm, so "
                          "a so'm to'lov's edit modal shows 1000 where the operator typed "
                          "12 000 000. Saqlash without touching anything re-runs convert_pair "
                          "on 1000 AS SO'M (crm/forms.py:161): hamkor-tolovlari.xlsx goes "
                          "12 000 000 so'm → 1 000 so'm on the first save and 0.08 so'm on "
                          "the second, and Hamkorga to'langan on the report collapses with "
                          "it. Unlike a formset row a plain ModelForm has no unchanged-row "
                          "skip, so ONE Saqlash is enough. Fix as ReturnForm does "
                          "(crm/forms.py:907-917, the one form that already swaps the so'm "
                          "side in): seed initial[typed_field] from the so'm column "
                          "when instance.is_som.",
                   strict=False)
def test_re_saving_a_som_hamkor_tolovi_unchanged_keeps_the_exported_figures(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    assert admin_client.post("/supplier-payments/new/", {
        "contract": contract.pk, "date": "2026-07-02", "currency": "uzs",
        "amount": "12000000", "exchange_rate": "12000", "commission_percent": "",
        "method": "cash", "fee_percent": "0", "note": "",
    }).status_code == 302
    payment = SupplierPayment.objects.get()
    before = _rows(admin_client.get(SUP_PAYS))[0]

    url = f"/supplier-payments/{payment.pk}/edit/"
    # what the modal actually shows: the dollar twin, under a So'm currency picker
    shown = _resubmit_payload(admin_client, url)
    assert shown["currency"] == "uzs"
    assert Decimal(shown["amount"]) == payment.amount == Decimal("1000.00")

    for _ in range(2):      # twice: the drift compounds
        assert admin_client.post(
            url, _resubmit_payload(admin_client, url)).status_code in (302, 200, 204)

    after = _rows(admin_client.get(SUP_PAYS))[0]
    assert after["Valyuta"] == "So'm"
    assert _dec(after["Hamkorga (so'm)"]) == _dec(before["Hamkorga (so'm)"])
    assert _dec(after["Hamkorga ($)"]) == _dec(before["Hamkorga ($)"])


@pytest.mark.xfail(reason="UPHELD — same known defect, PriceEntryFormMixin side, reaching "
                          "sotuvlar.xlsx. SaleForm (crm/forms.py:810) renders the derived "
                          "USD narx 1.1700 in the box beside a Valyuta select reading So'm. "
                          "One untouched Saqlash reads 1.17 as so'm: Sotuv narx (so'm) goes "
                          "14 040 → 1.17 and Jami (so'm) 14 040 000 → 1 170; a second goes "
                          "to 0.00, so Foyda ($) turns hugely negative and the report's "
                          "Sotuvdan foyda with it. Fix: seed initial['price'] from price_uzs "
                          "when instance.is_som, as ReturnForm (crm/forms.py:917) does.",
                   strict=False)
def test_re_saving_a_som_sotuv_unchanged_keeps_the_exported_figures(admin_client, db):
    lot = _lot(_contract(kg="100000", price="1.00"), kg="50000")
    assert admin_client.post(f"/sales/new/?lot={lot.pk}", {
        "customer": _customer().pk, "kg": "1000", "currency": "uzs",
        "price": "14040", "exchange_rate": "12000", "date": "2026-07-18",
        "debt_deadline": "", "note": "",
    }).status_code == 302
    sale = Sale.objects.get()
    before = _rows(admin_client.get(SALES))[0]

    url = f"/sales/{sale.pk}/edit/"
    shown = _resubmit_payload(admin_client, url)
    assert shown["currency"] == "uzs"
    assert Decimal(shown["price"]) == sale.price == Decimal("1.1700")

    for _ in range(2):
        assert admin_client.post(
            url, _resubmit_payload(admin_client, url)).status_code in (302, 200, 204)

    after = _rows(admin_client.get(SALES))[0]
    assert after["Valyuta"] == "So'm"
    assert _dec(after["Sotuv narx (so'm)"]) == _dec(before["Sotuv narx (so'm)"])
    assert _dec(after["Jami (so'm)"]) == _dec(before["Jami (so'm)"])


# rewritten: the original body called an undefined _rendered() and only ever raised
# NameError, so it was evidence of nothing. Rebuilt on _resubmit_payload, the same
# helper the two so'm tests above use, so it is the true control — identical round
# trip, dollar row, no drift.
def test_a_dollar_sotuv_re_saved_unchanged_does_not_move(admin_client, db):
    """The control for the two above: a dollar row survives the same round trip,
    which is what pins the defect to the so'm side rather than to the edit view."""
    lot = _lot(_contract(kg="100000", price="1.00"), kg="50000")
    assert admin_client.post(f"/sales/new/?lot={lot.pk}", {
        "customer": _customer().pk, "kg": "1000", "currency": "usd",
        "price": "1.17", "exchange_rate": "12000", "date": "2026-07-18",
        "debt_deadline": "", "note": "",
    }).status_code == 302
    sale = Sale.objects.get()
    before = _rows(admin_client.get(SALES))[0]

    url = f"/sales/{sale.pk}/edit/"
    shown = _resubmit_payload(admin_client, url)
    assert shown["currency"] == "usd"
    assert Decimal(shown["price"]) == sale.price == Decimal("1.1700")

    for _ in range(2):
        assert admin_client.post(
            url, _resubmit_payload(admin_client, url)).status_code in (302, 200, 204)

    after = _rows(admin_client.get(SALES))[0]
    assert after["Valyuta"] == "Dollar"
    assert _dec(after["Sotuv narx ($)"]) == _dec(before["Sotuv narx ($)"])
    assert _dec(after["Sotuv narx (so'm)"]) == _dec(before["Sotuv narx (so'm)"])
    assert _dec(after["Jami (so'm)"]) == _dec(before["Jami (so'm)"])


def test_exporting_twice_in_a_row_returns_identical_numbers(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    lot = _lot(contract, kg="50000")
    _sale(_customer(), lot, kg="1000", price="1.5000", currency="uzs",
          rate="13500", price_uzs="20250")
    SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.UZS,
        amount=Decimal("740.74"), amount_uzs=Decimal("10000000.00"),
        exchange_rate=Decimal("13500"), method="cash")

    for url in ALL_EXPORTS:
        first = _rows(admin_client.get(url))
        second = _rows(admin_client.get(url))
        assert first == second, url


# ── boundaries ───────────────────────────────────────────────────────────────

def test_an_empty_result_still_exports_a_header_only_file(admin_client, db):
    for url in ALL_EXPORTS:
        ws = _sheet(admin_client.get(url))
        assert _header(ws), url
        assert list(ws.iter_rows(min_row=2)) == [], url


# Regression guard. Was an xfail documenting a 500 on a malformed ?from/?to; it passes
# since _date_param (crm/views.py) drops anything that is not a real ISO date.
# Kept as a test so the crash cannot come back.
def test_a_malformed_date_filter_does_not_crash_the_report_or_the_exports(admin_client, db):
    _contract(kg="10000")
    bad = {"from": "kecha", "to": ""}
    got = {url: _status_or_error(admin_client, url, bad)
           for url in ("/reports/",) + ALL_EXPORTS}
    broken = {u: s for u, s in got.items() if s not in (200, 302, 400)}
    assert broken == {}, broken


@pytest.mark.xfail(reason="UPHELD — same unvalidated-querystring root cause as the "
                          "malformed-date crash, other fields. crm/views.py:2271 and :2283 "
                          "pass ?partner= and ?status= to filter(partner_id=...) / "
                          "filter(status_id=...) as raw strings, so /reports/?partner=abc "
                          "raises ValueError → HTTP 500 — 10 of the 12 URL/param pairs "
                          "probed here break, again sparing only qarzdorlar.xlsx. Fix: "
                          "coerce both to int in "
                          "_report_filters (crm/views.py:2251) and ignore what will not "
                          "coerce.",
                   strict=False)
def test_a_malformed_partner_or_status_filter_does_not_crash(admin_client, db):
    _contract(kg="10000")
    got = {}
    for bad in ({"partner": "abc"}, {"status": "abc"}):
        for url in ("/reports/",) + ALL_EXPORTS:
            got[(url, tuple(bad))] = _status_or_error(admin_client, url, bad)
    broken = {k: v for k, v in got.items() if v not in (200, 302, 400)}
    assert broken == {}, broken


def test_an_unknown_marka_filter_empties_the_report_rather_than_erroring(admin_client, db):
    """The boundary that IS handled: a brand string matching nothing is a plain
    empty result, not an error — the control that shows the two crashes above are
    about type coercion, not about unknown filter values."""
    _contract(kg="10000")
    q = {"brand": "YO'Q-MARKA"}
    assert admin_client.get("/reports/", q).status_code == 200
    for url in ALL_EXPORTS:
        assert _sheet(admin_client.get(url, q)) is not None


def test_a_tiny_kurs_keeps_the_typed_som_figure_exact(admin_client, db):
    """kurs = 1 so'm/$ is nonsense but enterable. The typed so'm side must still be
    the one that reaches the file, undivided."""
    partner = _partner()
    # the kurs is inherited now, not typed — patched to the nonsense figure the
    # probe is about
    with patch("crm.forms.latest_exchange_rate", return_value=Decimal("1")):
        assert admin_client.post("/contracts/new/", {
            "partner": partner.pk, "currency": "uzs", "created": "2026-07-01",
            "note": "", "planned_trucks": "1",
            **line_data({"brand": "LLDPE", "kg": "1000", "price": "9768"}),
        }).status_code == 302

    row = _rows(admin_client.get(CONTRACTS))[0]
    assert _dec(row["Narx (so'm)"]) == Decimal("9768")
    assert _dec(row["Narx ($)"]) == Decimal("9768")


def test_deleting_one_hamkor_tolovi_leaves_the_other_rows_untouched(admin_client, db):
    contract = _contract(kg="100000", price="1.00")
    keep = SupplierPayment.objects.create(
        contract=contract, date="2026-07-02", currency=Currency.UZS,
        amount=Decimal("740.74"), amount_uzs=Decimal("10000000.00"),
        exchange_rate=Decimal("13500"), method="cash")
    drop = SupplierPayment.objects.create(
        contract=contract, date="2026-07-03", currency=Currency.USD,
        amount=Decimal("500.00"), amount_uzs=Decimal("6000000.00"),
        exchange_rate=Decimal("12000"), method="cash")
    before = next(r for r in _rows(admin_client.get(SUP_PAYS))
                  if str(r["Sana"])[:10] == "2026-07-02")

    assert admin_client.post(f"/supplier-payments/{drop.pk}/delete/").status_code in (302, 204)

    rows = _rows(admin_client.get(SUP_PAYS))
    assert len(rows) == 1
    assert rows[0] == before
    screen = admin_client.get("/reports/").context
    assert screen["hamkorga_tolangan"] == keep.amount
    assert screen["hamkorga_tolangan_uzs"] == keep.amount_uzs
