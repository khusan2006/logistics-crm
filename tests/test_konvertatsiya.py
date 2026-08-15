"""Konvertatsiya — pul kassa ichida bir joydan ikkinchisiga o'tishi.

The till holds three heaps in two currencies, and the operator moves money between
them all day: naqd so'm sold for naqd dollars, cash walked into the bank, a karta
drawn out over the counter. Until this row existed there was nowhere to say so, and
the only ways to make the screen agree with the safe were a fake kirim and a fake
chiqim — two lies that also inflated both daftar.

What every test here is really guarding: an exchange moves money BETWEEN heaps and
into or out of NEITHER the business nor a daftar.
"""
from datetime import date, timedelta
from io import BytesIO
from decimal import Decimal

import openpyxl
from django.utils import timezone

from crm.models import (
    AuditLog,
    Currency,
    Customer,
    CustomerPayment,
    Konvertatsiya,
    PayMethod,
    kassa_cash_by_currency,
    kassa_cash_by_method,
)


def _customer(name="Alisher"):
    return Customer.objects.create(name=name, phone="1")


def _in(amount, method=PayMethod.CASH, currency=Currency.USD, day=10, rate="12000"):
    """Money in the till to change — a mijoz to'lov, the ordinary way it gets there."""
    amount = Decimal(amount)
    return CustomerPayment.objects.create(
        customer=_customer(), date=date(2026, 7, day), amount=amount,
        currency=currency, exchange_rate=Decimal(rate),
        amount_uzs=amount * Decimal(rate) if currency == Currency.USD else amount,
        method=method)


def _swap(from_amount="12000000", to_amount="1000", day=11,
          from_method=PayMethod.CASH, from_currency=Currency.UZS,
          to_method=PayMethod.CASH, to_currency=Currency.USD, **kw):
    """12 000 000 so'm naqd sold for $1 000 naqd, unless a test says otherwise."""
    return Konvertatsiya.objects.create(
        date=date(2026, 7, day),
        from_method=from_method, from_currency=from_currency,
        from_amount=Decimal(from_amount),
        to_method=to_method, to_currency=to_currency, to_amount=Decimal(to_amount),
        **kw)


def _heaps(rows=None):
    return dict(kassa_cash_by_currency(rows))


def _methods(client):
    return {m["label"]: dict(m["split_full"])
            for m in client.get("/kassa/?davr=all").context["cash_by_method"]}


def _form(**kw):
    """POST payload for the modal — a naqd so'm → naqd dollar swap by default."""
    data = {"date": "2026-07-11",
            "from_method": PayMethod.CASH, "from_currency": Currency.UZS,
            "from_amount": "12000000",
            "to_method": PayMethod.CASH, "to_currency": Currency.USD,
            "to_amount": "1000", "note": ""}
    data.update(kw)
    return data


class TestTheMoneyChangesHeaps:
    def test_som_sold_for_dollars_empties_one_pile_into_the_other(self, db):
        _in("12000000", currency=Currency.UZS)
        _swap()
        assert _heaps() == {Currency.USD: Decimal("1000.00")}

    def test_each_side_moves_in_full_and_neither_is_reconverted(self, db):
        """The kurs is already spent in the two figures. Applying it again — to the
        so'm side, or to the dollar one — is how a heap ends up holding money that was
        never in it."""
        _in("2000")                                   # $2 000 naqd
        _in("24000000", currency=Currency.UZS)        # 24 mln so'm naqd
        _swap(from_amount="1000", from_currency=Currency.USD,
              to_amount="12500000", to_currency=Currency.UZS)
        assert _heaps() == {Currency.USD: Decimal("1000.00"),
                            Currency.UZS: Decimal("36500000.00")}

    def test_the_usul_split_follows_it(self, admin_client, db):
        """The split this row exists for: naqd walked onto a karta changes nothing
        about how much there is and everything about which card it is on."""
        _in("5000000", currency=Currency.UZS)
        _swap(from_amount="3000000", from_currency=Currency.UZS,
              to_method=PayMethod.CARD, to_currency=Currency.UZS,
              to_amount="3000000")
        rows = _methods(admin_client)
        assert rows["Naqd"][Currency.UZS] == Decimal("2000000.00")
        assert rows["Karta"][Currency.UZS] == Decimal("3000000.00")

    def test_a_heap_can_be_moved_between_usul_without_changing_currency(self, db):
        _in("1000", method=PayMethod.TRANSFER)
        _swap(from_amount="400", from_currency=Currency.USD,
              from_method=PayMethod.TRANSFER,
              to_amount="400", to_currency=Currency.USD, to_method=PayMethod.CASH)
        by_method = {m["code"]: dict(m["split"]) for m in kassa_cash_by_method()}
        assert by_method["transfer"] == {Currency.USD: Decimal("600.00")}
        assert by_method["cash"] == {Currency.USD: Decimal("400.00")}

    def test_it_is_neither_a_kirim_nor_a_chiqim(self, admin_client, db):
        """Booked in the daftar it would be both at once, and both totals would count
        money that never crossed the door."""
        _in("12000000", currency=Currency.UZS)
        _swap()
        ctx = admin_client.get("/kassa/?davr=all").context
        assert list(ctx["outflow_page"].object_list) == []
        assert [r["kind"] for r in ctx["income_page"].object_list] == ["customer"]
        assert ctx["net_out"] == Decimal("0")

    def test_a_clean_exchange_leaves_the_blended_figure_alone(self, admin_client, db):
        """Dollars sold for so'm are the same money in another shape: the pair the
        waterfall closes on must not move by a tiyin."""
        _in("12000000", currency=Currency.UZS)
        before = admin_client.get("/kassa/?davr=all").context
        _swap()
        after = admin_client.get("/kassa/?davr=all").context
        assert after["cash_total"] == before["cash_total"]
        assert after["cash_total_uzs"] == before["cash_total_uzs"]

    def test_what_the_move_cost_is_what_survives(self, admin_client, db):
        """A million so'm that lands on a karta as 990 000 left 10 000 with the bank.
        The heaps show both halves in full; this is the only figure that says the
        business is poorer for it."""
        _in("1000000", currency=Currency.UZS)
        _swap(from_amount="1000000", from_currency=Currency.UZS,
              to_amount="990000", to_currency=Currency.UZS, to_method=PayMethod.CARD)
        ctx = admin_client.get("/kassa/?davr=all").context
        assert ctx["cash_total_uzs"] == Decimal("990000.00")
        assert _heaps() == {Currency.UZS: Decimal("990000.00")}

    def test_the_kurs_is_the_one_the_deal_struck(self, db):
        """Never the day's rate: what the operator actually got for their dollars is
        the only figure this row can honestly print."""
        row = _swap(from_amount="1000", from_currency=Currency.USD,
                    to_amount="12750000", to_currency=Currency.UZS)
        assert row.deal_rate == Decimal("12750.00")
        row.refresh_from_db()
        assert row.exchange_rate == Decimal("12750.00")

    def test_a_same_currency_move_prints_no_kurs(self, db):
        """It struck none — the stored one was inherited and decides nothing."""
        row = _swap(from_amount="500", from_currency=Currency.USD,
                    to_amount="500", to_currency=Currency.USD,
                    to_method=PayMethod.CARD)
        assert row.crosses_currency is False
        assert row.deal_rate is None


class TestTheForm:
    def test_admin_can_change_money(self, admin_client, db):
        resp = admin_client.post("/konvertatsiya/new/", _form(note="dollar oldik"))
        assert resp.status_code in (200, 302), resp.content[:400]
        row = Konvertatsiya.objects.get()
        assert (row.from_amount, row.to_amount) == (Decimal("12000000.00"),
                                                    Decimal("1000.00"))
        assert row.exchange_rate == Decimal("12000.00")
        assert row.note == "dollar oldik"

    def test_it_records_who_entered_it(self, admin_client, db):
        admin_client.post("/konvertatsiya/new/", _form())
        assert Konvertatsiya.objects.get().created_by is not None

    def test_both_sides_land_in_the_audit_trail(self, admin_client, db):
        """Either sum alone is half a fact — it cannot say whether that money arrived
        or left."""
        admin_client.post("/konvertatsiya/new/", _form())
        entry = AuditLog.objects.filter(target_type="Konvertatsiya").first()
        assert "12 000 000 so'm" in entry.summary.replace("\xa0", " ")
        assert "$1 000" in entry.summary.replace("\xa0", " ")

    def test_the_kurs_box_is_a_calculator_and_the_sums_are_the_record(self, admin_client, db):
        """It is typed to work the second figure out — "bugun 12 700 dan" — and the
        valyutachi then rounds. What the row reports is what the money did: 12 000 000
        handed over for $950 is 12 631.58, whatever the box said on the way."""
        admin_client.post("/konvertatsiya/new/",
                          _form(exchange_rate="12700", to_amount="950"))
        assert Konvertatsiya.objects.get().exchange_rate == Decimal("12631.58")

    def test_a_blank_kurs_is_the_ordinary_case(self, admin_client, db):
        """The two sums say what it was, so the box is never demanded."""
        admin_client.post("/konvertatsiya/new/", _form(exchange_rate=""))
        assert Konvertatsiya.objects.get().exchange_rate == Decimal("12000.00")

    def test_a_same_currency_move_inherits_a_kurs_rather_than_asking(self, admin_client, db):
        """Naqd so'm onto a karta strikes no rate at all. The column still needs a
        value — it is what gives the row's twin figure one — so it takes the last kurs
        anybody typed rather than a number the operator would have to invent."""
        from crm.models import latest_exchange_rate
        admin_client.post("/konvertatsiya/new/",
                          _form(from_currency=Currency.UZS, from_amount="1000000",
                                to_currency=Currency.UZS, to_amount="990000",
                                to_method=PayMethod.CARD))
        row = Konvertatsiya.objects.get()
        assert row.deal_rate is None
        assert row.exchange_rate == latest_exchange_rate()

    def test_a_move_to_the_same_heap_is_refused(self, admin_client, db):
        """It takes money out of a place and puts the same money back: a row that
        means "nothing happened", and almost always a half-filled form."""
        admin_client.post("/konvertatsiya/new/",
                          _form(from_currency=Currency.USD, from_amount="500",
                                to_currency=Currency.USD, to_amount="500"))
        assert not Konvertatsiya.objects.exists()

    def test_a_zero_or_negative_side_is_refused(self, admin_client, db):
        admin_client.post("/konvertatsiya/new/", _form(to_amount="0"))
        admin_client.post("/konvertatsiya/new/", _form(from_amount="-5"))
        assert not Konvertatsiya.objects.exists()

    def test_tomorrow_is_refused(self, admin_client, db):
        """Money moves when it moves — see reject_future. A future-dated row would
        move a heap the page insists has not moved yet."""
        tomorrow = (timezone.localdate() + timedelta(days=1)).isoformat()
        admin_client.post("/konvertatsiya/new/", _form(date=tomorrow))
        assert not Konvertatsiya.objects.exists()

    def test_an_edit_restates_the_kurs_off_the_new_pair(self, admin_client, db):
        row = _swap()
        admin_client.post(f"/konvertatsiya/{row.pk}/edit/",
                          _form(to_amount="960"))
        row.refresh_from_db()
        assert row.to_amount == Decimal("960.00")
        assert row.exchange_rate == Decimal("12500.00")

    def test_deleting_puts_the_money_back_where_it_was(self, admin_client, db):
        _in("12000000", currency=Currency.UZS)
        row = _swap()
        admin_client.post(f"/konvertatsiya/{row.pk}/delete/")
        assert not Konvertatsiya.objects.exists()
        assert _heaps() == {Currency.UZS: Decimal("12000000.00")}

    def test_the_modals_render(self, admin_client, db):
        """A form whose template blows up 500s only when somebody opens it, long
        after the page linking to it was called done. Both ways in: the kassa opens
        these in a modal (X-Requested-With), and the same URL is a full page."""
        row = _swap()
        ajax = {"headers": {"X-Requested-With": "XMLHttpRequest"}}
        for url in ("/konvertatsiya/new/", f"/konvertatsiya/{row.pk}/edit/",
                    f"/konvertatsiya/{row.pk}/delete/"):
            assert admin_client.get(url).status_code == 200, url
            modal = admin_client.get(url, **ajax)
            assert modal.status_code == 200, url
        # The two halves the form is built around, and the kurs box between them with
        # the hooks its calculator runs on (base.html).
        html = admin_client.get("/konvertatsiya/new/", **ajax).content.decode()
        assert "Qayerdan chiqdi" in html and "Qayerga tushdi" in html
        for hook in ("data-swap-rate", "data-swap-from", "data-swap-to",
                     "data-swap-from-currency", "data-swap-to-currency"):
            assert hook in html, hook

    def test_translator_forbidden(self, translator_client, db):
        row = _swap()
        assert translator_client.get("/konvertatsiya/new/").status_code == 403
        assert translator_client.post(f"/konvertatsiya/{row.pk}/delete/").status_code == 403
        assert Konvertatsiya.objects.exists()


class TestTheKassaScreen:
    def _rows(self, client, qs="?davr=all"):
        return list(client.get("/kassa/" + qs).context["exchange_page"].object_list)

    def test_the_table_reads_both_sides_on_one_line(self, admin_client, db):
        _swap(note="Sardor akadan")
        row = self._rows(admin_client)[0]
        assert (row["from_label"], row["from_amount"]) == ("Naqd", Decimal("12000000.00"))
        assert (row["to_label"], row["to_amount"]) == ("Naqd", Decimal("1000.00"))
        assert row["rate"] == Decimal("12000.00")
        html = admin_client.get("/kassa/?davr=all").content.decode()
        assert "Konvertatsiya" in html and "Sardor akadan" in html

    def test_the_davr_narrows_it(self, admin_client, db):
        _swap(day=2)
        _swap(day=20)
        rows = self._rows(admin_client, "?from=2026-07-15&to=2026-07-31")
        assert [r["date"] for r in rows] == [date(2026, 7, 20)]

    def test_the_search_box_covers_it_too(self, admin_client, db):
        _swap(note="bozordan")
        _swap(day=12, note="bankdan")
        assert len(self._rows(admin_client, "?davr=all&q=bozordan")) == 1
        # …and by either sum, however the operator groups the digits.
        assert len(self._rows(admin_client, "?davr=all&q=12 000 000")) == 2
        assert len(self._rows(admin_client, "?davr=all&q=yoq-narsa")) == 0

    def test_an_empty_davr_draws_no_table(self, admin_client, db):
        """An empty third table on every visit is furniture, not information."""
        html = admin_client.get("/kassa/?davr=all").content.decode()
        assert "Konvertatsiya" in html            # the button still says it exists
        assert "kassa ichida pulning" not in html.casefold()

    def test_the_excel_file_carries_its_own_tab(self, admin_client, db):
        _swap()
        resp = admin_client.get("/kassa/export.xlsx", {"davr": "all"})
        book = openpyxl.load_workbook(BytesIO(resp.content))
        assert "Konvertatsiya" in book.sheetnames
        values = [c.value for c in next(book["Konvertatsiya"].iter_rows(min_row=2))]
        assert Decimal(str(values[3])) == Decimal("12000000")
        assert Decimal(str(values[6])) == Decimal("1000")
