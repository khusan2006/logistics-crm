from decimal import Decimal

from conftest import payment_rows
from crm.templatetags.crm_extras import NBSP

from crm.models import (
    Contract, ContractLine, Customer, Partner, Sale, Shipment, ShipmentLine, ShipmentStatus,
)


def _customer(name="Alisher Mebel"):
    return Customer.objects.create(name=name, phone="1", address="Toshkent")


def _lot(kg="10000", brand="LLDPE", contract_price="1.00"):
    partner = Partner.objects.create(name="Pars", phone="1", city="T")
    contract = Contract.objects.create(partner=partner, created="2026-07-01")
    contract_line = ContractLine.objects.create(
        contract=contract, brand=brand, kg=Decimal(kg), price=Decimal(contract_price))
    _ship_obj = Shipment.objects.create(contract=contract, status=ShipmentStatus.arrival(), sent="2026-07-05", eta="2026-07-15", arrived="2026-07-16", transport="01A111AA", container="MSCU-1")
    _ship_obj_line = ShipmentLine.objects.create(
        shipment=_ship_obj, contract_line=contract.lines.first(), kg=Decimal(kg))
    return _ship_obj_line


def _sale(customer, lot, kg, price, date, debt_deadline=None):
    return Sale.objects.create(
        customer=customer, line=lot, kg=Decimal(kg), price=Decimal(price),
        date=date, debt_deadline=debt_deadline,
    )


def test_customer_with_unpaid_sale_appears_with_correct_total(admin_client, db):
    customer = _customer()
    lot = _lot()
    _sale(customer, lot, "1000", "1.60", "2026-07-17")

    html = admin_client.get("/debts/").content.decode()
    assert customer.name in html
    assert f"1{NBSP}600" in html


def test_fully_paid_customer_not_in_debt_list(admin_client, db):
    customer = _customer(name="Paid Customer")
    lot = _lot()
    sale = _sale(customer, lot, "1000", "1.60", "2026-07-17")
    admin_client.post("/customer-payments/new/", payment_rows(
        {"amount": "1600"}, customer=customer, date="2026-07-18"))
    sale.refresh_from_db()
    assert sale.remaining == Decimal("0")

    html = admin_client.get("/debts/").content.decode()
    assert "Paid Customer" not in html


def test_advance_customer_not_in_debt_list(admin_client, db):
    customer = _customer(name="Avans Customer")
    admin_client.post("/customer-payments/new/", payment_rows(
        {"amount": "500"}, customer=customer, date="2026-07-18"))
    customer.refresh_from_db()
    assert customer.balance < 0

    html = admin_client.get("/debts/").content.decode()
    assert "Avans Customer" not in html


def test_debt_customer_lists_outstanding_sales_and_excludes_paid(admin_client, db):
    customer = _customer()
    lot = _lot()
    unpaid = _sale(customer, lot, "1000", "1.60", "2026-07-17")
    paid = _sale(customer, lot, "500", "1.00", "2026-07-16")
    admin_client.post("/customer-payments/new/", payment_rows(
        {"amount": "500"}, customer=customer, date="2026-07-18"))
    paid.refresh_from_db()
    unpaid.refresh_from_db()
    assert paid.remaining == Decimal("0")
    assert unpaid.remaining > Decimal("0")

    resp = admin_client.get(f"/debts/{customer.pk}/")
    assert resp.status_code == 200
    html = resp.content.decode()
    assert f"#{lot.pk}" in html
    # Exactly one outstanding sale row: the unpaid one is listed, the fully-paid
    # sale is excluded. (Dates render localized, so count rows rather than date strings.)
    assert html.count('class="row-actions"') == 1
    assert f"$1{NBSP}600" in html          # the outstanding sale's total is shown


def test_overdue_sale_shows_overdue_indicator(admin_client, db):
    customer = _customer(name="Overdue Customer")
    lot = _lot()
    _sale(customer, lot, "1000", "1.60", "2026-07-01", debt_deadline="2026-07-10")

    list_html = admin_client.get("/debts/").content.decode()
    assert "Overdue Customer" in list_html
    assert ">1<" in list_html or "muddati o'tgan" in list_html.lower() or "kechikkan" in list_html.lower()

    detail_html = admin_client.get(f"/debts/{customer.pk}/").content.decode()
    assert "muddati o'tgan" in detail_html.lower() or "kechikkan" in detail_html.lower()


def test_sale_without_muddat_is_due_the_day_it_was_sold(admin_client, db):
    """A blank To'lov muddati means "pay now", not "never due" — otherwise an unpaid
    sotuv sits outside every Qarzlar signal."""
    customer = _customer()
    lot = _lot()
    sale = _sale(customer, lot, "1000", "1.60", "2026-07-17", debt_deadline=None)

    sale.refresh_from_db()
    assert str(sale.debt_deadline) == "2026-07-17"


def test_due_customers_sort_above_bigger_debts_not_yet_owed(admin_client, db):
    """Whoever has to pay now leads the table, oldest muddat first — even when a
    mijoz further down owes far more on a muddat that has not arrived."""
    from django.utils import timezone
    from datetime import timedelta

    today = timezone.localdate()
    lot = _lot(kg="100000")
    small_and_due = _customer(name="AAA Bugun")
    big_but_later = _customer(name="ZZZ Keyin")
    oldest_due = _customer(name="MMM Eski")

    _sale(small_and_due, lot, "100", "1.00", str(today), debt_deadline=str(today))
    _sale(big_but_later, lot, "50000", "1.00", str(today),
          debt_deadline=str(today + timedelta(days=30)))
    _sale(oldest_due, lot, "200", "1.00", str(today - timedelta(days=10)),
          debt_deadline=str(today - timedelta(days=10)))

    html = admin_client.get("/debts/").content.decode()
    order = [html.index(c.name) for c in (oldest_due, small_and_due, big_but_later)]
    assert order == sorted(order), "due-first, oldest muddat at the top"
    assert "bugun to'lash kerak" in html
    assert "muddati o'tgan" in html


class TestTheQarzlarRow:
    """Qarzlar carries the same row as every other ro'yxat: search on the left,
    ‹ davr › and Excel on the right.

    It had none of them for a while, on the reasoning that a qarz is a current-state
    figure with no window to honour. But the table prints a MUDDAT, and the person on
    this screen is chasing money by the day it was due — so the davr reads against the
    muddat, and the search reads the name and telefon a debtor is looked up by."""

    def _names(self, admin_client, qs=""):
        rows = admin_client.get("/debts/" + qs).context["page"].object_list
        return [r["customer"].name for r in rows]

    def test_the_row_has_a_search_a_davr_and_an_excel(self, admin_client, db):
        html = admin_client.get("/debts/").content.decode()
        assert (html.index('class="listbar"')
                < html.index('class="searchbar')
                < html.index('class="daterange-bar"'))
        assert "/debts/export.xlsx" in html

    def test_the_search_reads_the_name_and_the_telefon(self, admin_client, db):
        lot = _lot()
        alisher = Customer.objects.create(name="Alisher Mebel", phone="901112233")
        bekzod = Customer.objects.create(name="Bekzod Plast", phone="935556677")
        _sale(alisher, lot, "100", "1.00", "2026-07-17")
        _sale(bekzod, lot, "100", "1.00", "2026-07-17")

        assert self._names(admin_client, "?q=alisher") == ["Alisher Mebel"]
        assert self._names(admin_client, "?q=9355") == ["Bekzod Plast"]

    def test_the_davr_narrows_by_the_muddat(self, admin_client, db):
        lot = _lot()
        july = _customer(name="Iyulda to'laydi")
        august = _customer(name="Avgustda to'laydi")
        _sale(july, lot, "100", "1.00", "2026-07-01", debt_deadline="2026-07-20")
        _sale(august, lot, "100", "1.00", "2026-07-01", debt_deadline="2026-08-20")

        assert self._names(admin_client, "?from=2026-07-01&to=2026-07-31") == ["Iyulda to'laydi"]
        assert self._names(admin_client, "?from=2026-08-01&to=2026-08-31") == ["Avgustda to'laydi"]
        assert sorted(self._names(admin_client)) == sorted([july.name, august.name])

    def test_a_muddat_that_has_not_arrived_yet_is_still_findable(self, admin_client, db):
        """The row shows the oldest muddat that has already come — but the window has
        to reach the ones still ahead, or "kimning puli kelasi oyda kerak" has no
        answer and the ‹ arrow › forward lands on an empty page every time."""
        from datetime import timedelta

        from django.utils import timezone
        lot = _lot()
        later = _customer(name="Keyin to'laydi")
        deadline = timezone.localdate() + timedelta(days=20)
        _sale(later, lot, "100", "1.00", "2026-07-01", debt_deadline=str(deadline))

        rows = admin_client.get("/debts/", {"from": str(deadline), "to": str(deadline)})
        listed = rows.context["page"].object_list
        assert [r["customer"].name for r in listed] == ["Keyin to'laydi"]
        # …and the row still says it is not due yet.
        assert listed[0]["earliest_due"] is None

    def test_a_qarz_with_no_muddat_at_all_stays_out_of_a_chosen_window(self, admin_client, db):
        """It has no place on a calendar. With the filter off it is there as always."""
        lot = _lot()
        undated = _customer(name="Muddatsiz")
        _sale(undated, lot, "100", "1.00", "2026-07-01")
        Sale.objects.filter(customer=undated).update(debt_deadline=None)

        assert self._names(admin_client) == ["Muddatsiz"]
        assert self._names(admin_client, "?from=2026-07-01&to=2026-07-31") == []

    def test_the_excel_file_is_what_the_screen_is_showing(self, admin_client, db):
        """The promise every other Excel button on the app makes — the download is the
        searched, filtered list, not the whole table."""
        import openpyxl
        from io import BytesIO
        lot = _lot()
        alisher = Customer.objects.create(name="Alisher Mebel", phone="901112233")
        bekzod = Customer.objects.create(name="Bekzod Plast", phone="935556677")
        _sale(alisher, lot, "100", "1.00", "2026-07-17")
        _sale(bekzod, lot, "100", "1.00", "2026-07-17")

        sheet = openpyxl.load_workbook(
            BytesIO(admin_client.get("/debts/export.xlsx?q=alisher").content)).active
        names = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
        assert names == ["Alisher Mebel"]

    def test_the_hisobotlar_link_is_still_the_whole_table(self, admin_client, db):
        """That one is a report, not a screen — narrowing it by somebody else's
        search would make the two buttons quietly mean different things."""
        import openpyxl
        from io import BytesIO
        lot = _lot()
        alisher = Customer.objects.create(name="Alisher Mebel", phone="901112233")
        bekzod = Customer.objects.create(name="Bekzod Plast", phone="935556677")
        _sale(alisher, lot, "100", "1.00", "2026-07-17")
        _sale(bekzod, lot, "100", "1.00", "2026-07-17")

        sheet = openpyxl.load_workbook(BytesIO(
            admin_client.get("/reports/export/debts.xlsx?q=alisher").content)).active
        names = sorted(row[0] for row in sheet.iter_rows(min_row=2, values_only=True))
        assert names == ["Alisher Mebel", "Bekzod Plast"]


def test_translator_forbidden(translator_client, db):
    customer = _customer()
    assert translator_client.get("/debts/").status_code == 403
    assert translator_client.get(f"/debts/{customer.pk}/").status_code == 403


# ── Mijoz tarixi ─────────────────────────────────────────────────────────────

def _history(admin_client, customer):
    """The Mijoz tarixi rows of a mijoz's page, as (sana, voqea, tafsilot, summa)."""
    import re

    html = admin_client.get(f"/debts/{customer.pk}/").content.decode()
    assert "Mijoz tarixi" in html, "tarix bo'limi yo'q"
    section = html.split("Mijoz tarixi")[1]
    rows = re.findall(r"<tr>(.*?)</tr>", section, re.S)[1:]   # drop the header row

    def text(cell):
        # Collapses ASCII whitespace only: the NBSP thousands separator is the house
        # convention being asserted, so it must survive being read back out.
        return re.sub(r"[ \t\r\n]+", " ", re.sub(r"<[^>]+>", "", cell)).strip()

    return [tuple(text(cell) for cell in re.findall(r"<td.*?>(.*?)</td>", row, re.S))
            for row in rows]


def test_the_history_carries_all_four_kinds_of_event_newest_first(admin_client, db):
    """Sotuv, to'lov, qaytarish va bron bitta vaqt chizig'ida — the page answers
    "what has gone on with this mijoz", and that question is chronological."""
    from crm.models import CustomerPayment, Reservation, Return

    customer = _customer()
    lot = _lot()
    sale = _sale(customer, lot, "1000", "1.00", "2026-07-10")
    CustomerPayment.objects.create(customer=customer, date="2026-07-12",
                                   amount=Decimal("300.00"), method="cash")
    Return.objects.create(sale=sale, kg=Decimal("100"), price=Decimal("1.00"),
                          date="2026-07-14")
    Reservation.objects.create(customer=customer, brand="LLDPE", kg=Decimal("500"),
                               price=Decimal("1.20"))

    rows = _history(admin_client, customer)
    labels = [r[1] for r in rows]
    assert "Sotuv" in labels
    assert "To&#x27;lov" in labels
    assert "Qaytarish" in labels
    assert any(label.startswith("Bron") for label in labels)
    # newest first — parsed, because d.m.Y does not sort as text
    from datetime import datetime

    dates = [datetime.strptime(r[0], "%d.%m.%Y").date() for r in rows]
    assert dates == sorted(dates, reverse=True)


def test_each_history_row_is_drawn_in_the_currency_it_moved_in(admin_client, db):
    """A dollar sotuv and a so'm to'lov stay two separate facts — nothing on this
    timeline is converted into the other currency."""
    from crm.models import Currency, CustomerPayment

    customer = _customer()
    _sale(customer, _lot(), "1000", "1.00", "2026-07-10")          # $1 000
    CustomerPayment.objects.create(
        customer=customer, date="2026-07-12", amount=Decimal("500.00"),
        amount_uzs=Decimal("6000000.00"), exchange_rate=Decimal("12000"),
        currency=Currency.UZS, method="cash")

    amounts = {r[1]: r[3] for r in _history(admin_client, customer)}
    assert amounts["Sotuv"] == f"$1{NBSP}000"
    assert amounts["To&#x27;lov"] == f"6{NBSP}000{NBSP}000 so&#x27;m"


def test_a_bron_with_no_narx_agreed_says_so_rather_than_inventing_one(admin_client, db):
    from crm.models import Reservation

    customer = _customer()
    Reservation.objects.create(customer=customer, brand="LLDPE", kg=Decimal("500"))
    row = _history(admin_client, customer)[0]
    assert row[3] == "kelishilmagan"


def test_a_mijoz_with_no_dealings_gets_an_empty_history_not_a_crash(admin_client, db):
    customer = _customer()
    html = admin_client.get(f"/debts/{customer.pk}/").content.decode()
    assert "Bu mijoz bilan hali hech qanday amaliyot bo'lmagan" in html


def test_the_mijozlar_list_links_the_name_to_that_mijoz_page(admin_client, db):
    """Tahrirlash is the pencil beside the name, so the name itself is free to open
    the page that actually says something about the mijoz."""
    customer = _customer()
    html = admin_client.get("/customers/").content.decode()
    assert f'href="/debts/{customer.pk}/">{customer.name}</a>' in html
