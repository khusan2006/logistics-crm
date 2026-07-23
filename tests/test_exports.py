from decimal import Decimal
from io import BytesIO

import openpyxl

from crm.models import (
    Contract, ContractLine, Customer, CustomerPayment, Partner, Sale, Shipment, ShipmentLine, ShipmentStatus, SupplierPayment,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EXPORT_URLS = {
    "/reports/export/contracts.xlsx": [
        "Kelishuv", "Sana", "Hamkor", "Marka", "Kg", "Narx", "Jami", "Yuborilgan kg",
        "To'langan", "Qarz",
    ],
    "/reports/export/supplier-payments.xlsx": ["Sana", "Kelishuv", "Hamkor", "Summa", "Usul"],
    "/reports/export/shipments.xlsx": [
        "Yuk ID", "Kelishuv", "Hamkor", "Marka", "Kg", "Holat", "Jo'natilgan", "Reja kelish",
        "Yetib kelgan", "Transport", "Konteyner",
    ],
    "/reports/export/sales.xlsx": [
        "Sana", "Mijoz", "Lot ID", "Marka", "Kg", "Tan narx", "Sotuv narx", "Jami", "Foyda", "Qoldiq",
    ],
    "/reports/export/debts.xlsx": ["Mijoz", "Telefon", "Jami savdo", "To'langan", "Qarz"],
}


def _contract(partner=None, brand="LLDPE", created="2026-07-01"):
    partner = partner or Partner.objects.create(name="Pars", phone="1", city="T")
    _contract_obj = Contract.objects.create(partner=partner, created=created, deadline="2026-08-01")
    _contract_obj_line = ContractLine.objects.create(
        contract=_contract_obj, brand=brand, kg=Decimal("1000"), price=Decimal("1"))
    return _contract_obj


def _arrived_shipment(contract, kg=Decimal("500"), eta="2026-07-15", arrived="2026-07-16"):
    _ship_obj = Shipment.objects.create(contract=contract, status=ShipmentStatus.arrival(), sent="2026-07-05", eta=eta, arrived=arrived, transport="01A111AA", container="MSCU-1")
    _ship_obj_line = ShipmentLine.objects.create(
        shipment=_ship_obj, contract_line=contract.lines.first(), kg=kg)
    return _ship_obj


def _customer(name="Alisher Mebel"):
    return Customer.objects.create(name=name, phone="998901112233", address="Toshkent")


def _sale(customer, shipment, kg=Decimal("100"), price=Decimal("2"), cost_price=Decimal("1"),
          date="2026-07-17"):
    return Sale.objects.create(
        customer=customer, line=shipment.lines.first(), kg=kg, price=price, cost_price=cost_price, date=date,
    )


def _seed():
    partner = Partner.objects.create(name="Pars", phone="1", city="T")
    contract = _contract(partner=partner)
    shipment = _arrived_shipment(contract)
    customer = _customer()
    sale = _sale(customer, shipment)
    sup_pay = SupplierPayment.objects.create(
        contract=contract, date="2026-07-11", amount=Decimal("200.00"), method="cash",
    )
    CustomerPayment.objects.create(
        customer=customer, date="2026-07-17", amount=Decimal("50.00"), method="cash",
    )
    return {
        "partner": partner, "contract": contract, "shipment": shipment,
        "customer": customer, "sale": sale, "sup_pay": sup_pay,
    }


def _load(resp):
    return openpyxl.load_workbook(BytesIO(resp.content))


def test_exports_return_xlsx_with_header_and_rows(admin_client, db):
    _seed()
    for url, headers in EXPORT_URLS.items():
        resp = admin_client.get(url)
        assert resp.status_code == 200, url
        assert resp["Content-Type"] == XLSX_MIME, url
        assert "attachment" in resp["Content-Disposition"], url
        assert ".xlsx" in resp["Content-Disposition"], url
        assert resp.content, url

        wb = _load(resp)
        ws = wb.active
        header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header_row == headers, url
        data_rows = list(ws.iter_rows(min_row=2))
        assert len(data_rows) == 1, url


def test_partner_filter_excludes_other_partner_contract(admin_client, db):
    data = _seed()
    other_partner = Partner.objects.create(name="Kaveh", phone="2", city="S")
    _contract(partner=other_partner, created="2026-07-05")

    resp = admin_client.get("/reports/export/contracts.xlsx", {"partner": data["partner"].pk})
    assert resp.status_code == 200
    wb = _load(resp)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2))
    assert len(data_rows) == 1
    hamkor_names = [row[2].value for row in data_rows]
    assert "Kaveh" not in hamkor_names
    assert "Pars" in hamkor_names


def test_translator_forbidden_on_every_export(translator_client, db):
    for url in EXPORT_URLS:
        assert translator_client.get(url).status_code == 403, url
