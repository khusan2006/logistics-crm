"""The Excel button every ro'yxat carries.

The point of these is one sentence: the file holds what the SCREEN was showing. An
export that quietly ignores the page's search, filters or davr is worse than no
export — the figures in it look official and answer a question nobody asked. So each
button goes through the same filter helper its page does, and these tests hold the two
together.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from conftest import make_contract, make_lot, make_shipment
from crm.models import (
    AuditLog,
    Customer,
    CustomerPayment,
    Partner,
    Sale,
    ShipmentStatus,
    SupplierPayment,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Every list's button, and the file it downloads.
LIST_EXPORTS = {
    "/contracts/export.xlsx": "kelishuvlar.xlsx",
    "/shipments/export.xlsx": "yuklar.xlsx",
    "/sales/export.xlsx": "sotuvlar.xlsx",
    "/customer-payments/export.xlsx": "mijoz-tolovlari.xlsx",
    "/supplier-payments/export.xlsx": "hamkor-tolovlari.xlsx",
    "/ombor/export.xlsx": "ombor.xlsx",
    "/audit/export.xlsx": "audit.xlsx",
    "/kassa/export.xlsx": "kassa.xlsx",
}


def _sheet(resp, index=0):
    return openpyxl.load_workbook(BytesIO(resp.content)).worksheets[index]


def _rows(resp, index=0):
    return list(_sheet(resp, index).iter_rows(min_row=2, values_only=True))


def _customer(name="Alisher"):
    return Customer.objects.create(name=name, phone="1")


@pytest.mark.parametrize("url,filename", LIST_EXPORTS.items())
def test_every_list_downloads_a_real_workbook(admin_client, db, url, filename):
    resp = admin_client.get(url)
    assert resp.status_code == 200
    assert resp["Content-Type"] == XLSX_MIME
    assert f'filename="{filename}"' in resp["Content-Disposition"]
    ws = _sheet(resp)
    assert [c.value for c in ws[1]], "sarlavha qatori bo'sh"
    # Read like a report, not like a raw dump: header bold and frozen in place.
    assert ws["A1"].font.bold and ws.freeze_panes == "A2"


def test_the_file_holds_the_window_the_page_was_showing(admin_client, db):
    customer = _customer()
    lot = make_lot(kg="5000")
    Sale.objects.create(customer=customer, line=lot, kg=Decimal("10"),
                        price=Decimal("1"), date=date(2026, 7, 10))
    Sale.objects.create(customer=customer, line=lot, kg=Decimal("10"),
                        price=Decimal("1"), date=date(2026, 8, 10))

    july = admin_client.get("/sales/export.xlsx",
                            {"from": "2026-07-01", "to": "2026-07-31"})
    assert len(_rows(july)) == 1
    assert _rows(july)[0][0].date() == date(2026, 7, 10)
    assert len(_rows(admin_client.get("/sales/export.xlsx"))) == 2


def test_the_file_holds_the_search_the_page_was_showing(admin_client, db):
    pars = Partner.objects.create(name="Pars", phone="1", city="T")
    kaveh = Partner.objects.create(name="Kaveh", phone="2", city="S")
    make_contract(partner=pars, brand="LLDPE")
    make_contract(partner=kaveh, brand="HDPE")

    found = _rows(admin_client.get("/contracts/export.xlsx", {"q": "Kaveh"}))
    assert [row[2] for row in found] == ["Kaveh"]


def test_the_yuklar_file_follows_the_hammasi_toggle(admin_client, db):
    """The default view is the loads still moving; Hammasi adds the arrived ones. The
    file is whichever set the button was pressed on."""
    # Built with lines: the file is one row per PRODUCT, as the hisobotlar export is.
    make_shipment(kg="400", eta="2026-07-20")
    make_shipment(kg="400", eta="2026-07-10", arrived="2026-07-12",
                  status=ShipmentStatus.arrival())

    moving = admin_client.get("/shipments/export.xlsx")
    everything = admin_client.get("/shipments/export.xlsx", {"all": "1"})
    assert len(_rows(moving)) == 1
    assert len(_rows(everything)) == 2


def test_the_kassa_file_is_two_tabs(admin_client, db):
    """Kirim and Chiqim are read against each other, so they arrive in one file."""
    customer = _customer()
    CustomerPayment.objects.create(customer=customer, date=date(2026, 7, 5),
                                   amount=Decimal("100"))
    SupplierPayment.objects.create(contract=make_contract(), date=date(2026, 7, 6),
                                   amount=Decimal("40"))

    wb = openpyxl.load_workbook(BytesIO(admin_client.get("/kassa/export.xlsx?davr=all").content))
    assert [ws.title for ws in wb.worksheets] == ["Kirim", "Chiqim"]
    assert len(list(wb["Kirim"].iter_rows(min_row=2))) == 1
    assert len(list(wb["Chiqim"].iter_rows(min_row=2))) == 1


def test_money_lands_as_a_number_that_excel_can_add_up(admin_client, db):
    """A figure formatted into a string is a figure nobody can sum — the one thing a
    spreadsheet is for."""
    SupplierPayment.objects.create(contract=make_contract(), date=date(2026, 7, 6),
                                   amount=Decimal("1234.50"), method="cash")

    ws = _sheet(admin_client.get("/supplier-payments/export.xlsx"))
    headers = [c.value for c in ws[1]]
    cell = ws.cell(row=2, column=headers.index("Hamkorga ($)") + 1)
    assert isinstance(cell.value, (int, float, Decimal))
    assert cell.number_format == "#,##0.00"
    # A sana is a real date, not text that sorts alphabetically.
    sana = ws.cell(row=2, column=headers.index("Sana") + 1)
    assert sana.number_format == "DD.MM.YYYY"


def test_kg_keeps_its_three_decimals(admin_client, db):
    """Rounded to two, 0.125 kg of a marka disappears into the row above it."""
    make_contract(kg="1000")
    ws = _sheet(admin_client.get("/contracts/export.xlsx", {"state": ""}))
    headers = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=headers.index("Kg") + 1).number_format == "#,##0.000"


def test_the_audit_file_is_cut_to_the_same_window_as_the_page(admin_client, db):
    AuditLog.objects.create(action=AuditLog.Action.CREATE, target_type="Sotuv",
                            summary="iyulda")
    entry = AuditLog.objects.get()
    AuditLog.objects.filter(pk=entry.pk).update(created_at="2026-07-31 18:30:00+05:00")

    inside = admin_client.get("/audit/export.xlsx",
                              {"from": "2026-07-01", "to": "2026-07-31"})
    outside = admin_client.get("/audit/export.xlsx",
                               {"from": "2026-08-01", "to": "2026-08-31"})
    assert len(_rows(inside)) == 1
    assert len(_rows(outside)) == 0


def test_a_translator_gets_only_the_two_lists_they_can_read(translator_client, db):
    """The same rule as the pages themselves: kelishuvlar and yuklar are readable
    without the money columns, and everything else is admin-only."""
    assert translator_client.get("/contracts/export.xlsx").status_code == 200
    assert translator_client.get("/shipments/export.xlsx").status_code == 200
    for url in ("/sales/export.xlsx", "/customer-payments/export.xlsx",
                "/supplier-payments/export.xlsx", "/kassa/export.xlsx",
                "/audit/export.xlsx"):
        assert translator_client.get(url).status_code == 403, url


def test_a_skladchi_can_take_the_ombor_but_not_the_kassa(skladchi_client, db):
    assert skladchi_client.get("/ombor/export.xlsx").status_code == 200
    assert skladchi_client.get("/kassa/export.xlsx?davr=all").status_code == 403
